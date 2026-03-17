from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

wb = Workbook()

# ── Shared styles ──
ARIAL = Font(name='Arial', size=10)
ARIAL_BOLD = Font(name='Arial', size=10, bold=True)
HEADER_FILL = PatternFill('solid', fgColor='D9D9D9')
YELLOW_FILL = PatternFill('solid', fgColor='FFFFCC')
BLUE_FILL = PatternFill('solid', fgColor='CCE5FF')
GREEN_FILL = PatternFill('solid', fgColor='CCFFCC')
THIN_BORDER = Border(
    bottom=Side(style='thin', color='BFBFBF')
)
CURRENCY_FMT = '$#,##0'
DATE_FMT = 'YYYY-MM-DD'
TIME_FMT = 'HH:mm'
TIMESTAMP_FMT = 'YYYY-MM-DD HH:mm'

# ═══════════════════════════════════════════
# TAB 1: Timesheet_Submissions
# ═══════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Timesheet_Submissions'

headers = [
    'Submission Timestamp',  # A
    'Staff Name',            # B
    'Phone Number',          # C
    'Date of Work',          # D
    'Work Venue',            # E
    'Basic Rate ($)',        # F
    'Start Time',            # G
    'End Time',              # H
    'Notes',                 # I
    'Project No.',           # J
    'OT Compensation ($)',   # K
    'Final Rate ($)',        # L
    'Status',                # M
]

col_widths = {
    'A': 22, 'B': 18, 'C': 20, 'D': 14, 'E': 22,
    'F': 14, 'G': 12, 'H': 12, 'I': 28, 'J': 14,
    'K': 20, 'L': 14, 'M': 12,
}

for col_letter, w in col_widths.items():
    ws1.column_dimensions[col_letter].width = w

# Header row
for ci, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=ci, value=h)
    cell.font = ARIAL_BOLD
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER

ws1.freeze_panes = 'A2'

# Phone Number column as text
ws1.column_dimensions['C'].number_format = '@'

# ── Sample data ──
from datetime import datetime, time

sample_rows = [
    {
        'A': datetime(2026, 2, 20, 10, 0),
        'B': 'Alice Chen',
        'C': '+852 1234 5678',
        'D': datetime(2026, 2, 20),
        'E': 'HKCEC Hall 3',
        'F': 600,
        'G': time(9, 0),
        'H': time(18, 0),
        'I': '',
        'J': '',
        'K': '',
        'M': 'Pending',
    },
    {
        'A': datetime(2026, 2, 20, 10, 15),
        'B': 'Bob Smith',
        'C': '+852 9876 5432',
        'D': datetime(2026, 2, 20),
        'E': 'HKCEC Hall 3',
        'F': 600,
        'G': time(9, 0),
        'H': time(22, 0),
        'I': 'Stayed for OT till 10pm',
        'J': '',
        'K': '',
        'M': 'Pending',
    },
    {
        'A': datetime(2026, 2, 20, 11, 0),
        'B': 'Alice Chen',
        'C': '+852 1234 5678',
        'D': datetime(2026, 2, 21),
        'E': 'City Hall',
        'F': 800,
        'G': time(10, 0),
        'H': time(19, 0),
        'I': '',
        'J': 'P2026-003',
        'K': 200,
        'M': 'Approved',
    },
]

for ri, row_data in enumerate(sample_rows, 2):
    for col_letter, value in row_data.items():
        cell = ws1[f'{col_letter}{ri}']
        cell.value = value
        cell.font = ARIAL

    # Formula in L
    ws1[f'L{ri}'] = f'=IF(K{ri}="",F{ri},F{ri}+K{ri})'
    ws1[f'L{ri}'].font = ARIAL

# ── Number formats per column (apply to rows 2–1000 for future data) ──
for r in range(2, 1001):
    ws1[f'A{r}'].number_format = TIMESTAMP_FMT
    ws1[f'C{r}'].number_format = '@'  # text
    ws1[f'D{r}'].number_format = DATE_FMT
    ws1[f'F{r}'].number_format = CURRENCY_FMT
    ws1[f'G{r}'].number_format = TIME_FMT
    ws1[f'H{r}'].number_format = TIME_FMT
    ws1[f'K{r}'].number_format = CURRENCY_FMT
    ws1[f'L{r}'].number_format = CURRENCY_FMT
    # Default font for all cells
    for c in 'ABCDEFGHIJKLM':
        if ws1[f'{c}{r}'].font == Font():
            ws1[f'{c}{r}'].font = ARIAL

# ── Data validation on Status (M) ──
dv = DataValidation(
    type='list',
    formula1='"Pending,Approved,Paid"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle='Invalid Status',
    error='Please select Pending, Approved, or Paid.',
)
dv.sqref = 'M2:M1000'
ws1.add_data_validation(dv)

# ── Conditional formatting on Status (M2:M1000) ──
ws1.conditional_formatting.add(
    'M2:M1000',
    CellIsRule(operator='equal', formula=['"Pending"'], fill=YELLOW_FILL)
)
ws1.conditional_formatting.add(
    'M2:M1000',
    CellIsRule(operator='equal', formula=['"Approved"'], fill=BLUE_FILL)
)
ws1.conditional_formatting.add(
    'M2:M1000',
    CellIsRule(operator='equal', formula=['"Paid"'], fill=GREEN_FILL)
)

# ═══════════════════════════════════════════
# TAB 2: Staff_Directory
# ═══════════════════════════════════════════
ws2 = wb.create_sheet('Staff_Directory')

dir_headers = ['Staff Name', 'Phone Number', 'Usual Roles', 'Notes']
dir_widths = [20, 22, 24, 30]

for ci, (h, w) in enumerate(zip(dir_headers, dir_widths), 1):
    cell = ws2.cell(row=1, column=ci, value=h)
    cell.font = ARIAL_BOLD
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN_BORDER
    ws2.column_dimensions[get_column_letter(ci)].width = w

ws2.freeze_panes = 'A2'

# Set Phone Number column as text
for r in range(2, 201):
    ws2[f'B{r}'].number_format = '@'
    for c in 'ABCD':
        ws2[f'{c}{r}'].font = ARIAL

# ═══════════════════════════════════════════
# TAB 3: Dashboard
# ═══════════════════════════════════════════
ws3 = wb.create_sheet('Dashboard')

TITLE_FONT = Font(name='Arial', size=14, bold=True)
LABEL_FONT = Font(name='Arial', size=11, bold=True)
VALUE_FONT = Font(name='Arial', size=11)

ws3.column_dimensions['A'].width = 32
ws3.column_dimensions['B'].width = 18

ws3['A1'] = 'Timesheet Dashboard'
ws3['A1'].font = TITLE_FONT

dashboard_rows = [
    ('Total Pending Submissions', '=COUNTIF(Timesheet_Submissions!M:M,"Pending")', None),
    ('Total Approved (Unpaid)', '=COUNTIF(Timesheet_Submissions!M:M,"Approved")', None),
    ('Total Paid', '=COUNTIF(Timesheet_Submissions!M:M,"Paid")', None),
    ('Total Payroll (Approved)', '=SUMIF(Timesheet_Submissions!M:M,"Approved",Timesheet_Submissions!L:L)', CURRENCY_FMT),
]

for i, (label, formula, fmt) in enumerate(dashboard_rows):
    r = i + 3  # start from row 3
    ws3.cell(row=r, column=1, value=label).font = LABEL_FONT
    val_cell = ws3.cell(row=r, column=2)
    val_cell.value = formula
    val_cell.font = VALUE_FONT
    val_cell.alignment = Alignment(horizontal='center')
    if fmt:
        val_cell.number_format = fmt

ws3.freeze_panes = 'A2'

# ── Save ──
OUTPUT = '/Volumes/Mic Backup/TS HR Roster/deliverables/timesheet_template.xlsx'
wb.save(OUTPUT)
print(f'Saved to {OUTPUT}')
